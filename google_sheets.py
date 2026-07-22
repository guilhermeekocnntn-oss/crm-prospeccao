import os
import sys
import io
import json
import time
import unicodedata
import gspread

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_PATH = os.path.join(BASE_DIR, "credentials.json")

# Suporte ao diretório temporário do Vercel (/tmp)
if os.environ.get("VERCEL"):
    CACHE_FILE = "/tmp/leads_cache.json"
else:
    CACHE_FILE = os.path.join(BASE_DIR, "leads_cache.json")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

ESTADOS_MAP = {
    "AC": "Acre", "AL": "Alagoas", "AP": "Amapá", "AM": "Amazonas",
    "BA": "Bahia", "CE": "Ceará", "DF": "Distrito Federal", "ES": "Espírito Santo",
    "GO": "Goiás", "MA": "Maranhão", "MT": "Mato Grosso", "MS": "Mato Grosso do Sul",
    "MG": "Minas Gerais", "PA": "Pará", "PB": "Paraíba", "PR": "Paraná",
    "PE": "Pernambuco", "PI": "Piauí", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte",
    "RS": "Rio Grande do Sul", "RO": "Rondônia", "RR": "Roraima", "SC": "Santa Catarina",
    "SP": "São Paulo", "SE": "Sergipe", "TO": "Tocantins"
}

folder_cache = {}
DRIVE_CACHE_FILENAME = "crm_leads_cache.json"

def normalizar(texto):
    if not texto:
        return ""
    nfkd = unicodedata.normalize('NFD', str(texto))
    return "".join([c for c in nfkd if not unicodedata.combining(c)]).lower().strip()

def criar_servicos():
    """Conecta ao Google usando GOOGLE_CREDENTIALS do Vercel ou credentials.json local"""
    creds_json = (
        os.environ.get("GOOGLE_CREDENTIALS") or 
        os.environ.get("GOOGLE_CREDENTIALS_JSON") or 
        os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    )
    
    creds = None
    if creds_json:
        try:
            creds_info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_info, scopes=SCOPES)
        except Exception as e:
            print(f"Erro ao decodificar JSON das credenciais na variável de ambiente: {e}")
    
    if not creds and os.path.exists(CREDENTIALS_PATH):
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        
    if not creds:
        print("⚠️ Credenciais do Google não encontradas no ambiente nem em arquivo.")
        return None, None

    try:
        gclient = gspread.authorize(creds)
        drive_service = build('drive', 'v3', credentials=creds)
        return gclient, drive_service
    except Exception as e:
        print(f"Erro ao inicializar serviços do Google: {e}")
        return None, None

def adicionar_novo_lead_no_drive(dados_lead):
    """
    Localiza a planilha no Drive pela macrorregião, insere o lead na aba da cidade
    e atualiza o cache local sem apagar os leads existentes.
    """
    gclient, drive_service = criar_servicos()
    if not drive_service:
        raise Exception("Serviços do Google Drive indisponíveis.")

    macroregiao = dados_lead.get('macroregiao', '').strip()
    
    # Se a cidade/aba vier vazia, define "Geral" como padrão
    cidade_aba = dados_lead.get('aba', '').strip() or dados_lead.get('cidade', '').strip() or 'Geral'
    dados_lead['aba'] = cidade_aba
    dados_lead['cidade'] = cidade_aba

    query = "(mimeType='application/vnd.google-apps.spreadsheet' or mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') and trashed=false"
    results = drive_service.files().list(
        q=query, 
        fields="files(id, name, mimeType)", 
        supportsAllDrives=True, 
        includeItemsFromAllDrives=True
    ).execute()
    
    files = results.get('files', [])
    target_file = None

    for f in files:
        if macroregiao and (normalizar(macroregiao) in normalizar(f['name']) or normalizar(f['name']) in normalizar(macroregiao)):
            target_file = f
            break

    if not target_file and files:
        target_file = files[0]

    if not target_file:
        raise Exception("Nenhuma planilha correspondente foi encontrada no Google Drive.")

    sheet_id = target_file['id']
    mime_type = target_file.get('mimeType', '')
    linha_inserida = None

    nova_linha = [
        dados_lead.get('empresa', ''),
        dados_lead.get('tipo', ''),
        dados_lead.get('bairro', ''),
        dados_lead.get('telefone', ''),
        dados_lead.get('decisor', ''),
        dados_lead.get('instagram_site', ''),
        dados_lead.get('marca_propria', ''),
        dados_lead.get('potencial', 'Médio'),
        dados_lead.get('status', 'A Ligar (Novo)'),
        dados_lead.get('data_ultimo', ''),
        dados_lead.get('data_retorno', ''),
        dados_lead.get('resumo', '')
    ]

    if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        request = drive_service.files().get_media(fileId=sheet_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)

        wb = openpyxl.load_workbook(fh)
        
        aba_encontrada = None
        for sname in wb.sheetnames:
            if normalizar(sname) == normalizar(cidade_aba):
                aba_encontrada = sname
                break

        if not aba_encontrada:
            aba_encontrada = cidade_aba
            ws = wb.create_sheet(title=aba_encontrada)
            ws.append(["Nome da Empresa", "Tipo", "Bairro", "Telefone", "Decisor", "Instagram/Site", "Marca Própria", "Potencial", "Status", "Data Último", "Data Retorno", "Resumo"])
        else:
            ws = wb[aba_encontrada]

        ws.append(nova_linha)
        linha_inserida = ws.max_row

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        drive_service.files().update(fileId=sheet_id, media_body=media, supportsAllDrives=True).execute()

    else:
        sheet = gclient.open_by_key(sheet_id)
        worksheet = None
        
        for ws in sheet.worksheets():
            if normalizar(ws.title) == normalizar(cidade_aba):
                worksheet = ws
                break

        if not worksheet:
            worksheet = sheet.add_worksheet(title=cidade_aba, rows="100", cols="20")
            worksheet.append_row(["Nome da Empresa", "Tipo", "Bairro", "Telefone", "Decisor", "Instagram/Site", "Marca Própria", "Potencial", "Status", "Data Último", "Data Retorno", "Resumo"])

        worksheet.append_row(nova_linha)
        linha_inserida = len(worksheet.get_all_values())

    dados_lead['sheet_id'] = sheet_id
    dados_lead['linha_id'] = linha_inserida

    # Atualiza o cache local
    leads_cache = []
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                leads_cache = json.load(f)
        except Exception:
            leads_cache = []

    leads_cache.append(dados_lead)

    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(leads_cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ Aviso: Não foi possível salvar em arquivo local: {e}")

    salvar_cache_no_drive(leads_cache)

    return dados_lead

def salvar_cache_no_drive(todos_leads):
    try:
        _, drive_service = criar_servicos()
        if not drive_service:
            return False

        conteudo = json.dumps(todos_leads, ensure_ascii=False, indent=2).encode("utf-8")
        media = MediaIoBaseUpload(io.BytesIO(conteudo), mimetype="application/json", resumable=False)

        results = drive_service.files().list(
            q=f"name='{DRIVE_CACHE_FILENAME}' and trashed=false",
            fields="files(id, name)",
            spaces="drive"
        ).execute()
        arquivos = results.get("files", [])

        if arquivos:
            file_id = arquivos[0]["id"]
            drive_service.files().update(fileId=file_id, media_body=media).execute()
        else:
            metadata = {"name": DRIVE_CACHE_FILENAME, "mimeType": "application/json"}
            drive_service.files().create(body=metadata, media_body=media, fields="id").execute()
        return True
    except Exception as e:
        print(f"⚠️ Erro ao salvar cache no Drive: {e}")
        return False

def carregar_cache_do_drive():
    try:
        _, drive_service = criar_servicos()
        if not drive_service:
            return None

        results = drive_service.files().list(
            q=f"name='{DRIVE_CACHE_FILENAME}' and trashed=false",
            fields="files(id, name)",
            spaces="drive"
        ).execute()
        arquivos = results.get("files", [])

        if not arquivos:
            return None

        file_id = arquivos[0]["id"]
        request = drive_service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return json.loads(fh.read().decode("utf-8"))
    except Exception as e:
        print(f"⚠️ Erro ao carregar cache do Drive: {e}")
        return None

def obter_info_pasta(drive_service, folder_id):
    if not folder_id:
        return "", []
    if folder_id in folder_cache:
        return folder_cache[folder_id]['name'], folder_cache[folder_id]['parents']
    
    try:
        res = drive_service.files().get(
            fileId=folder_id, 
            fields='name, parents', 
            supportsAllDrives=True
        ).execute()
        name = res.get('name', '')
        parents = res.get('parents', [])
        folder_cache[folder_id] = {'name': name, 'parents': parents}
        return name, parents
    except Exception:
        return "", []

def identificar_estado_da_planilha(drive_service, parents):
    if not parents:
        return "OUTROS", "Geral", "Sem Pasta"
    
    curr_id = parents[0]
    visited = set()
    
    while curr_id and curr_id not in visited:
        visited.add(curr_id)
        folder_name, parent_parents = obter_info_pasta(drive_service, curr_id)
        if not folder_name:
            break
            
        norm_folder = normalizar(folder_name)
        
        for sigla, nome in ESTADOS_MAP.items():
            norm_nome = normalizar(nome)
            norm_sigla = normalizar(sigla)
            
            if norm_nome in norm_folder or norm_folder in norm_nome or norm_sigla == norm_folder:
                return sigla, nome, folder_name
                
        curr_id = parent_parents[0] if parent_parents else None
        
    return "OUTROS", "Geral", "Geral"

def extrair_valor(row, chaves_possiveis):
    for key in row.keys():
        norm_key = normalizar(key)
        for chave in chaves_possiveis:
            if normalizar(chave) in norm_key:
                val = row[key]
                return str(val).strip() if val is not None else ""
    return ""

def processar_item_planilha(item):
    gclient, drive_service = criar_servicos()
    if not drive_service:
        return []

    sheet_id = item['id']
    nome_planilha = item['name']
    mime_type = item.get('mimeType', '')
    parents = item.get('parents', [])
    
    sigla_uf, nome_uf, pasta_origem = identificar_estado_da_planilha(drive_service, parents)
    leads_item = []

    max_tentativas = 3
    for tentativa in range(1, max_tentativas + 1):
        try:
            if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                request = drive_service.files().get_media(fileId=sheet_id)
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                fh.seek(0)
                
                wb = openpyxl.load_workbook(fh, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows or len(rows) < 2:
                        continue
                    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
                    
                    for r_idx, row_values in enumerate(rows[1:], start=2):
                        row = {headers[i]: row_values[i] if i < len(row_values) and row_values[i] is not None else "" for i in range(len(headers))}
                        empresa = extrair_valor(row, ["Nome da Empresa", "Empresa", "Razao Social", "Nome"])
                        if empresa:
                            leads_item.append({
                                "sheet_id": sheet_id,
                                "linha_id": r_idx,
                                "uf": sigla_uf,
                                "uf_nome": nome_uf,
                                "pasta": pasta_origem,
                                "macroregiao": nome_planilha,
                                "aba": sheet_name,
                                "cidade": sheet_name.split("_")[0] if "_" in sheet_name else sheet_name,
                                "empresa": empresa,
                                "tipo": extrair_valor(row, ["Tipo", "Categoria"]),
                                "bairro": extrair_valor(row, ["Bairro", "Regiao", "Bairro / Região"]),
                                "telefone": extrair_valor(row, ["Telefone", "WhatsApp", "Contato", "Telefone / WhatsApp"]),
                                "decisor": extrair_valor(row, ["Decisor", "Nome do Decisor", "Socio"]),
                                "instagram_site": extrair_valor(row, ["Instagram", "Site", "Redes"]),
                                "marca_propria": extrair_valor(row, ["Marca Propria", "Possui Marca"]),
                                "potencial": extrair_valor(row, ["Potencial", "Potencial do Cliente"]) or "Médio",
                                "status": extrair_valor(row, ["Status", "Status do Contato"]) or "A Ligar (Novo)",
                                "data_ultimo": extrair_valor(row, ["Data Ultimo", "Ultimo Contato"]),
                                "data_retorno": extrair_valor(row, ["Data de Retorno", "Retorno"]),
                                "resumo": extrair_valor(row, ["Resumo", "Objeção", "Observacao", "Notas"])
                            })
            else:
                sheet = gclient.open_by_key(sheet_id)
                for worksheet in sheet.worksheets():
                    aba_title = worksheet.title
                    try:
                        registros = worksheet.get_all_records()
                    except Exception:
                        continue
                    for index, row in enumerate(registros, start=2):
                        empresa = extrair_valor(row, ["Nome da Empresa", "Empresa", "Razao Social", "Nome"])
                        if empresa:
                            leads_item.append({
                                "sheet_id": sheet_id,
                                "linha_id": index,
                                "uf": sigla_uf,
                                "uf_nome": nome_uf,
                                "pasta": pasta_origem,
                                "macroregiao": nome_planilha,
                                "aba": aba_title,
                                "cidade": aba_title.split("_")[0] if "_" in aba_title else aba_title,
                                "empresa": empresa,
                                "tipo": extrair_valor(row, ["Tipo", "Categoria"]),
                                "bairro": extrair_valor(row, ["Bairro", "Regiao", "Bairro / Região"]),
                                "telefone": extrair_valor(row, ["Telefone", "WhatsApp", "Contato", "Telefone / WhatsApp"]),
                                "decisor": extrair_valor(row, ["Decisor", "Nome do Decisor", "Socio"]),
                                "instagram_site": extrair_valor(row, ["Instagram", "Site", "Redes"]),
                                "marca_propria": extrair_valor(row, ["Marca Propria", "Possui Marca"]),
                                "potencial": extrair_valor(row, ["Potencial", "Potencial do Cliente"]) or "Médio",
                                "status": extrair_valor(row, ["Status", "Status do Contato"]) or "A Ligar (Novo)",
                                "data_ultimo": extrair_valor(row, ["Data Ultimo", "Ultimo Contato"]),
                                "data_retorno": extrair_valor(row, ["Data de Retorno", "Retorno"]),
                                "resumo": extrair_valor(row, ["Resumo", "Objeção", "Observacao", "Notas"])
                            })
            break

        except Exception as e:
            if tentativa == max_tentativas:
                print(f" ⚠️ Falha ao processar '{nome_planilha}': {e}")
            else:
                time.sleep(1)
                
    return leads_item

def obter_todos_leads_eko(force_refresh=False):
    if not force_refresh and os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                todos_leads = json.load(f)
                return todos_leads
        except Exception:
            pass

    if not force_refresh:
        leads_cache_drive = carregar_cache_do_drive()
        if leads_cache_drive is not None:
            try:
                with open(CACHE_FILE, 'w', encoding='utf-8') as f:
                    json.dump(leads_cache_drive, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            return leads_cache_drive

    gclient, drive_service = criar_servicos()
    if not drive_service:
        return []
    
    planilhas = []
    page_token = None
    query = "(mimeType='application/vnd.google-apps.spreadsheet' or mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') and trashed=false"
    
    try:
        while True:
            results = drive_service.files().list(
                q=query,
                fields="nextPageToken, files(id, name, mimeType, parents)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                pageSize=1000,
                pageToken=page_token
            ).execute()
            
            planilhas.extend(results.get('files', []))
            page_token = results.get('nextPageToken', None)
            if not page_token:
                break
    except Exception as e:
        print(f" ❌ ERRO ao listar arquivos do Drive: {e}")
        return []
    
    todos_leads = []
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(processar_item_planilha, item) for item in planilhas]
        for future in as_completed(futures):
            resultado = future.result()
            todos_leads.extend(resultado)

    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(todos_leads, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f" ⚠️ Erro ao salvar cache local: {e}")

    salvar_cache_no_drive(todos_leads)

    return todos_leads

def atualizar_lead_no_cache(sheet_id, nome_aba, linha, coluna_nome, novo_valor):
    if not os.path.exists(CACHE_FILE):
        return
    try:
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            leads_cache = json.load(f)
        for lead in leads_cache:
            if (str(lead.get('sheet_id')) == str(sheet_id)
                    and str(lead.get('aba')) == str(nome_aba)
                    and str(lead.get('linha_id')) == str(linha)):
                lead[coluna_nome] = novo_valor
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(leads_cache, f, ensure_ascii=False, indent=2)
        salvar_cache_no_drive(leads_cache)
    except Exception as e:
        print(f"Erro ao atualizar cache: {e}")

def atualizar_lead_tempo_real(sheet_id, nome_aba, linha, coluna_nome, novo_valor):
    gclient, drive_service = criar_servicos()
    if not drive_service:
        return False
    
    colunas = {
        "empresa": 1, "tipo": 2, "bairro": 3, "telefone": 4,
        "decisor": 5, "instagram_site": 6, "marca_propria": 7,
        "potencial": 8, "status": 9, "data_ultimo": 10,
        "data_retorno": 11, "resumo": 12
    }
    
    col_idx = colunas.get(coluna_nome)
    if not col_idx:
        return False

    atualizar_lead_no_cache(sheet_id, nome_aba, linha, coluna_nome, novo_valor)

    try:
        res = drive_service.files().get(fileId=sheet_id, fields='mimeType', supportsAllDrives=True).execute()
        mime_type = res.get('mimeType', '')

        if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            request = drive_service.files().get_media(fileId=sheet_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            fh.seek(0)

            wb = openpyxl.load_workbook(fh)
            if nome_aba in wb.sheetnames:
                ws = wb[nome_aba]
                ws.cell(row=int(linha), column=int(col_idx), value=str(novo_valor))
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
                drive_service.files().update(fileId=sheet_id, media_body=media, supportsAllDrives=True).execute()
                return True
        else:
            sheet = gclient.open_by_key(sheet_id)
            worksheet = sheet.worksheet(nome_aba)
            worksheet.update_cell(int(linha), col_idx, str(novo_valor))
            return True
    except Exception as e:
        print(f"❌ Erro ao atualizar lead no Drive: {e}")
        return False
