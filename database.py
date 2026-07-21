import sqlite3
import json
import gzip
import urllib.request
import os
import openpyxl

if os.environ.get("VERCEL"):
    DB_NAME = "/tmp/prospeccao.db"
else:
    DB_NAME = "prospeccao.db"

def get_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            uf TEXT NOT NULL,
            mesorregiao TEXT,
            cidade TEXT NOT NULL,
            empresa TEXT NOT NULL,
            tipo TEXT,
            bairro TEXT,
            telefone TEXT,
            decisor TEXT,
            instagram_site TEXT,
            marca_propria TEXT,
            potencial TEXT,
            status TEXT DEFAULT 'A Ligar (Novo)',
            data_ultimo_contato TEXT,
            data_retorno TEXT,
            resumo_conversa TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def buscar_cidades_ibge(uf):
    url = f"https://servicodados.ibge.gov.br/api/v1/localidades/estados/{uf}/municipios?ordenacao=nome"
    headers = {'User-Agent': 'Mozilla/5.0'}
    req = urllib.request.Request(url, headers=headers)
    
    with urllib.request.urlopen(req) as response:
        conteudo_bruto = response.read()
        if response.info().get('Content-Encoding') == 'gzip' or conteudo_bruto[:2] == b'\x1f\x8b':
            conteudo_bruto = gzip.decompress(conteudo_bruto)
        dados = json.loads(conteudo_bruto.decode('utf-8'))
        
    resultado = []
    for item in dados:
        cidade = item['nome']
        try:
            meso = item['microrregiao']['mesorregiao']['nome']
        except (KeyError, TypeError):
            meso = "Geral"
        resultado.append({"cidade": cidade, "mesorregiao": meso})
        
    return resultado

def importar_planilhas_pasta(pasta_origem):
    """Varre todas as planilhas da pasta e importa os registros para o banco de dados."""
    if not pasta_origem or not os.path.exists(pasta_origem):
        print(f"⚠️ Pasta não encontrada: {pasta_origem}")
        return 0

    init_db()
    conn = get_connection()
    cursor = conn.cursor()
    
    total_importados = 0
    
    for root_dir, _, files in os.walk(pasta_origem):
        for file in files:
            if (file.endswith(".xlsx") or file.endswith(".XLSX")) and not file.startswith("~$"):
                caminho_arquivo = os.path.join(root_dir, file)
                
                # Descobre UF pelo nome do arquivo (ex: SP_Mesorregiao_...)
                uf = file.split("_")[0].upper() if "_" in file else "SP"
                
                try:
                    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        cidade = sheet_name.split("_")[0] # Nome da aba é a cidade
                        
                        # Itera a partir da linha 2 (ignorando o cabeçalho)
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            empresa = row[0]
                            
                            # Se a coluna empresa estiver preenchida
                            if empresa and str(empresa).strip():
                                cursor.execute('''
                                    INSERT INTO leads (uf, cidade, empresa, tipo, bairro, telefone, decisor, instagram_site, marca_propria, potencial, status, data_ultimo_contato, data_retorno, resumo_conversa)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ''', (
                                    uf, cidade, str(empresa).strip(),
                                    str(row[1] or ''), str(row[2] or ''), str(row[3] or ''),
                                    str(row[4] or ''), str(row[5] or ''), str(row[6] or ''),
                                    str(row[7] or 'Médio'), str(row[8] or 'A Ligar (Novo)'),
                                    str(row[9] or ''), str(row[10] or ''), str(row[11] or '')
                                ))
                                total_importados += 1
                except Exception as e:
                    print(f"Erro ao ler arquivo {file}: {e}")
                    
    conn.commit()
    conn.close()
    return total_importados

def obter_todos_leads_db():
    """Retorna os leads armazenados no SQLite formatados no padrão do CRM."""
    init_db()
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM leads ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    
    leads = []
    for r in rows:
        lead = {
            "sheet_id": f"db_{r['id']}",
            "linha_id": r['id'],
            "uf": r['uf'],
            "uf_nome": r['uf'],
            "pasta": "SQLite Local",
            "macroregiao": r['cidade'],
            "aba": r['cidade'],
            "empresa": r['empresa'],
            "tipo": r['tipo'] or '',
            "bairro": r['bairro'] or '',
            "telefone": r['telefone'] or '',
            "decisor": r['decisor'] or '',
            "instagram_site": r['instagram_site'] or '',
            "marca_propria": r['marca_propria'] or '',
            "potencial": r['potencial'] or 'Médio',
            "status": r['status'] or 'A Ligar (Novo)',
            "data_ultimo": r['data_ultimo_contato'] or '',
            "data_retorno": r['data_retorno'] or '',
            "resumo": r['resumo_conversa'] or ''
        }
        leads.append(lead)
    return leads

def adicionar_lead_db(dados):
    """Insere um novo lead no banco de dados SQLite local com todos os campos da planilha."""
    init_db()
    conn = get_connection()
    cursor = conn.cursor()
    
    uf = (dados.get('uf') or 'SP').strip().upper()
    cidade = (dados.get('cidade') or 'Geral').strip()
    empresa = (dados.get('empresa') or '').strip()
    tipo = (dados.get('tipo') or '').strip()
    bairro = (dados.get('bairro') or '').strip()
    telefone = (dados.get('telefone') or '').strip()
    decisor = (dados.get('decisor') or '').strip()
    instagram = (dados.get('instagram_site') or '').strip()
    marca_propria = (dados.get('marca_propria') or '').strip()
    potencial = (dados.get('potencial') or 'Médio').strip()
    status = (dados.get('status') or 'A Ligar (Novo)').strip()
    data_ultimo = (dados.get('data_ultimo') or '').strip()
    data_retorno = (dados.get('data_retorno') or '').strip()
    resumo = (dados.get('resumo') or '').strip()

    cursor.execute('''
        INSERT INTO leads (uf, cidade, empresa, tipo, bairro, telefone, decisor, instagram_site, marca_propria, potencial, status, data_ultimo_contato, data_retorno, resumo_conversa)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (uf, cidade, empresa, tipo, bairro, telefone, decisor, instagram, marca_propria, potencial, status, data_ultimo, data_retorno, resumo))
    
    lead_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return {
        "sheet_id": f"db_{lead_id}",
        "linha_id": lead_id,
        "uf": uf,
        "uf_nome": uf,
        "pasta": "SQLite Local",
        "macroregiao": cidade,
        "aba": cidade,
        "cidade": cidade,
        "empresa": empresa,
        "tipo": tipo,
        "bairro": bairro,
        "telefone": telefone,
        "decisor": decisor,
        "instagram_site": instagram,
        "marca_propria": marca_propria,
        "potencial": potencial,
        "status": status,
        "data_ultimo": data_ultimo,
        "data_retorno": data_retorno,
        "resumo": resumo
    }

if __name__ == "__main__":
    init_db()
    print("Banco de dados pronto!")
